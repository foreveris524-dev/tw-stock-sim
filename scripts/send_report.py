import io, json, os, smtplib
from datetime import datetime, timezone, timedelta
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email import encoders

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

tz_tw = timezone(timedelta(hours=8))
today = datetime.now(tz_tw).strftime("%Y-%m-%d")

with open("portfolio.json") as f:
    p = json.load(f)
with open("futures_portfolio.json") as f:
    fut = json.load(f)

# ── 數值 ─────────────────────────────────────────────────────────────────────
cash          = p.get("cash", 0)
total_value   = p.get("total_value", 0)
holdings      = p.get("holdings", [])
initial_cap   = p.get("initial_capital", 1000000)
stock_log     = p.get("trade_log", [])
SKIP_ACTIONS  = {"RESET", "PLAN_BUY", "PLAN_SKIP", "PLAN_SHORT"}
today_trades  = [t for t in stock_log if t.get("date") == today and t.get("action") not in SKIP_ACTIONS]

position       = fut.get("position")
cum_pnl        = fut.get("cumulative_pnl", 0)
used_margin    = fut.get("used_margin", 0)
avail_margin   = fut.get("available_margin", 0)
init_margin    = fut.get("initial_margin", 200000)
fut_log        = fut.get("trade_log", [])
FUT_SKIP       = {"RESET", "PLAN_SHORT"}
today_fut      = [t for t in fut_log if t.get("date") == today and t.get("action") not in FUT_SKIP]
today_fut_pnl  = sum(t.get("pnl", 0) for t in today_fut)

stock_pnl   = total_value - initial_cap
net_assets  = total_value + avail_margin + cum_pnl
total_pnl   = net_assets - (initial_cap + init_margin)
margin_pct  = used_margin / init_margin * 100 if init_margin else 0

def sign(n):
    return "+" if n >= 0 else ""

# ── 郵件文字內容 ──────────────────────────────────────────────────────────────
lines = []
lines.append("📊 台股模擬操盤 每日報表")
lines.append(f"日期：{today}")
lines.append("")
lines.append("═══ 現貨帳戶 ═══")
lines.append(f"現金：${cash:,.0f}")
if holdings:
    lines.append("持股：")
    for h in holdings:
        pct  = h.get("value", 0) / total_value * 100 if total_value else 0
        upnl = h.get("unrealized_pnl", 0)
        upct = h.get("unrealized_pnl_pct", 0)
        lines.append(f"  {h['code']} {h.get('name','')}  均價${h.get('avg_price',0):,.1f}  "
                     f"{h.get('shares',0)}股  市值${h.get('value',0):,.0f}（{pct:.1f}%）")
        if upnl:
            lines.append(f"    未實現損益：{sign(upnl)}${upnl:,.0f}（{sign(upct)}{upct:.1f}%）")
else:
    lines.append("持股：無")
if today_trades:
    lines.append("今日操作：")
    for t in today_trades:
        lines.append(f"  {t['action']} {t.get('code','')} "
                     f"${t.get('price',0):,.0f} x {t.get('shares',0)}股")
else:
    lines.append("今日操作：觀望，無交易")
lines.append(f"帳戶總值：${total_value:,.0f}")
lines.append(f"相較起始：{sign(stock_pnl)}${stock_pnl:,.0f}")
lines.append("")
lines.append("═══ 期貨帳戶 ═══")
if position:
    d = "做多" if position.get("direction") == "LONG" else "做空"
    lines.append(f"目前持倉：{d} {position.get('contracts',0)}口 @ {position.get('entry_price',0):,}點")
    upnl = position.get("unrealized_pnl", 0)
    if upnl:
        lines.append(f"未實現損益：{sign(upnl)}${upnl:,.0f}")
else:
    lines.append("目前持倉：無")
if today_fut:
    lines.append("今日操作：")
    for t in today_fut:
        lines.append(f"  {t['action']}  損益：{sign(t.get('pnl',0))}${t.get('pnl',0):,.0f}")
else:
    lines.append("今日操作：觀望，無交易")
lines.append(f"今日損益：{sign(today_fut_pnl)}${today_fut_pnl:,.0f}")
lines.append(f"累計損益：{sign(cum_pnl)}${cum_pnl:,.0f}")
lines.append(f"保證金使用率：{margin_pct:.0f}%")
lines.append("")
lines.append("═══ 整體資產 ═══")
lines.append(f"總資產：${net_assets:,.0f}")
lines.append(f"整體損益：{sign(total_pnl)}${total_pnl:,.0f}")
if HAS_EXCEL:
    lines.append("")
    lines.append("（詳細操作紀錄請見附件 Excel）")

report = "\n".join(lines)
print(report)

# ── Excel 產生（記憶體） ────────────────────────────────────────────────────
def make_excel():
    C = dict(
        DARK="1E2329", HEADER="2B3139", BLUE="3B82F6",
        GREEN="10B981", RED="EF4444", YELLOW="F59E0B",
        LIGHT="252D36", TEXT="E5E7EB", MUTED="9CA3AF", WHITE="FFFFFF",
    )

    def fill(c):
        return PatternFill("solid", fgColor=c)

    def fnt(bold=False, color="E5E7EB", size=10):
        return Font(bold=bold, color=color, name="Calibri", size=size)

    def ctr():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def lft():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    def bdr():
        s = Side(border_style="thin", color="3A4351")
        return Border(left=s, right=s, top=s, bottom=s)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1：帳戶總覽 ──────────────────────────────────────────────────
    ws = wb.create_sheet("帳戶總覽")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C["BLUE"]
    for c, w in zip("ABCDE", [3, 22, 20, 20, 10]):
        ws.column_dimensions[c].width = w

    ws.merge_cells("B1:E1")
    t = ws["B1"]
    t.value = f"📊  台股模擬操盤  —  帳戶總覽  ({today})"
    t.font = Font(bold=True, color=C["WHITE"], name="Calibri", size=14)
    t.fill = fill(C["DARK"]); t.alignment = ctr()
    ws.row_dimensions[1].height = 36

    sections = [
        ("▌ 現貨帳戶", C["BLUE"], [
            ("可用現金",    f"${cash:,.0f}",         C["TEXT"]),
            ("持股市值",    f"${total_value-cash:,.0f}", C["TEXT"]),
            ("帳戶總值",    f"${total_value:,.0f}",   C["BLUE"]),
            ("持股檔數",    f"{len(holdings)} 檔",    C["TEXT"]),
        ]),
        ("▌ 期貨帳戶", C["YELLOW"], [
            ("目前持倉",    (f"{'空' if position and position.get('direction')=='SHORT' else '多'} "
                            f"{position.get('contracts',0)}口 @ {position.get('entry_price',0):,}"
                            if position else "無持倉"),         C["TEXT"]),
            ("可用保證金",  f"${avail_margin:,.0f}",  C["TEXT"]),
            ("保證金使用率",f"{margin_pct:.1f}%",
             C["RED"] if margin_pct > 70 else C["GREEN"]),
            ("累計損益",    f"${cum_pnl:+,.0f}",
             C["GREEN"] if cum_pnl >= 0 else C["RED"]),
        ]),
        ("▌ 整體資產", C["GREEN"], [
            ("整體總資產",  f"${net_assets:,.0f}",    C["WHITE"]),
            ("整體損益",    f"${total_pnl:+,.0f}",
             C["GREEN"] if total_pnl >= 0 else C["RED"]),
        ]),
    ]

    row = 3
    for sec_title, sec_color, rows in sections:
        ws.merge_cells(f"B{row}:E{row}")
        c = ws[f"B{row}"]
        c.value = sec_title
        c.font = Font(bold=True, color=sec_color, name="Calibri", size=11)
        c.fill = fill(C["DARK"]); c.alignment = lft()
        ws.row_dimensions[row].height = 28
        row += 1
        for label, val, vcol in rows:
            bg = C["HEADER"] if row % 2 == 0 else C["LIGHT"]
            for col in "BCDE":
                ws[f"{col}{row}"].fill = fill(bg)
                ws[f"{col}{row}"].border = bdr()
            ws[f"B{row}"].value = label
            ws[f"B{row}"].font = fnt(color=C["MUTED"])
            ws[f"B{row}"].alignment = lft()
            ws[f"C{row}"].value = val
            ws[f"C{row}"].font = fnt(bold=True, color=vcol, size=11)
            ws[f"C{row}"].alignment = lft()
            ws.row_dimensions[row].height = 24
            row += 1
        row += 1  # 段間空行，無格式

    # ── Sheet 2：現貨操作紀錄 ───────────────────────────────────────────────
    ws2 = wb.create_sheet("現貨操作紀錄")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = C["GREEN"]
    for ci, w in enumerate([3,12,8,14,8,10,10,12,52], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    ws2.merge_cells("B1:I1")
    t = ws2["B1"]
    t.value = "現貨操作紀錄"
    t.font = Font(bold=True, color=C["WHITE"], name="Calibri", size=13)
    t.fill = fill(C["DARK"]); t.alignment = ctr()
    ws2.row_dimensions[1].height = 30

    headers = ["日期","時間","動作","代號","價格","股數","金額","說明"]
    HCOL = {"BUY": C["GREEN"], "SELL": C["RED"], "ADJUST_LIMIT": C["YELLOW"],
            "PLAN_CANCEL": C["RED"], "OBSERVE": C["MUTED"], "RESET": C["MUTED"],
            "PLAN_BUY": "6EE7B7", "PLAN_SKIP": C["MUTED"]}
    for ci, h in enumerate(headers, 2):
        c = ws2.cell(row=2, column=ci)
        c.value = h; c.font = fnt(bold=True, color=C["WHITE"])
        c.fill = fill(C["GREEN"]); c.alignment = ctr(); c.border = bdr()
    ws2.row_dimensions[2].height = 22

    for ri, entry in enumerate(stock_log, 3):
        bg = C["LIGHT"] if ri % 2 == 0 else C["HEADER"]
        act = entry.get("action","")
        row_vals = [
            entry.get("date",""), entry.get("time",""), act,
            entry.get("code",""),
            entry.get("new_limit") or (entry.get("price") if entry.get("price") else ""),
            entry.get("shares","") or "",
            entry.get("amount","") or "",
            entry.get("reason",""),
        ]
        for ci, val in enumerate(row_vals, 2):
            cell = ws2.cell(row=ri, column=ci)
            cell.value = val; cell.fill = fill(bg); cell.border = bdr()
            if ci == 4:
                cell.font = fnt(bold=True, color=HCOL.get(act, C["TEXT"]))
                cell.alignment = ctr()
            elif ci in (5,6,7):
                cell.font = fnt(color=C["TEXT"])
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif ci == 9:
                cell.font = fnt(color=C["MUTED"], size=9); cell.alignment = lft()
            else:
                cell.font = fnt(color=C["TEXT"]); cell.alignment = ctr()
        ws2.row_dimensions[ri].height = 20

    # ── Sheet 3：期貨操作紀錄 ───────────────────────────────────────────────
    ws3 = wb.create_sheet("期貨操作紀錄")
    ws3.sheet_view.showGridLines = False
    ws3.sheet_properties.tabColor = C["YELLOW"]
    for ci, w in enumerate([3,12,8,14,8,8,10,12,52], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    ws3.merge_cells("B1:I1")
    t = ws3["B1"]
    t.value = "期貨操作紀錄（小台指 MTX）"
    t.font = Font(bold=True, color=C["WHITE"], name="Calibri", size=13)
    t.fill = fill(C["DARK"]); t.alignment = ctr()
    ws3.row_dimensions[1].height = 30

    fut_headers = ["日期","時間","動作","方向","口數","進場點/現價","損益(元)","說明"]
    FHCOL = {"SHORT": C["RED"], "LONG": C["GREEN"], "PLAN_SHORT": "FCA5A5",
             "PLAN_CANCEL": C["RED"], "OBSERVE": C["MUTED"], "RESET": C["MUTED"]}
    for ci, h in enumerate(fut_headers, 2):
        c = ws3.cell(row=2, column=ci)
        c.value = h; c.font = fnt(bold=True, color=C["WHITE"])
        c.fill = fill(C["YELLOW"]); c.alignment = ctr(); c.border = bdr()
    ws3.row_dimensions[2].height = 22

    for ri, entry in enumerate(fut_log, 3):
        bg = C["LIGHT"] if ri % 2 == 0 else C["HEADER"]
        act = entry.get("action","")
        direc = entry.get("direction","")
        pnl = entry.get("pnl", 0)
        row_vals = [
            entry.get("date",""), entry.get("time",""), act, direc or "—",
            entry.get("contracts","") or "—",
            entry.get("entry_price","") or entry.get("current_price","") or "—",
            pnl if pnl else "",
            entry.get("reason",""),
        ]
        for ci, val in enumerate(row_vals, 2):
            cell = ws3.cell(row=ri, column=ci)
            cell.value = val; cell.fill = fill(bg); cell.border = bdr()
            if ci == 4:
                cell.font = fnt(bold=True, color=FHCOL.get(act, C["TEXT"]))
                cell.alignment = ctr()
            elif ci == 5:
                dcol = C["RED"] if direc=="SHORT" else (C["GREEN"] if direc=="LONG" else C["MUTED"])
                cell.font = fnt(bold=True, color=dcol); cell.alignment = ctr()
            elif ci == 8:
                pcol = C["GREEN"] if (pnl or 0)>0 else (C["RED"] if (pnl or 0)<0 else C["MUTED"])
                cell.font = fnt(bold=True, color=pcol)
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif ci == 9:
                cell.font = fnt(color=C["MUTED"], size=9); cell.alignment = lft()
            else:
                cell.font = fnt(color=C["TEXT"]); cell.alignment = ctr()
        ws3.row_dimensions[ri].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ── 寄信 ─────────────────────────────────────────────────────────────────────
gmail_user     = "foreveris524@gmail.com"
gmail_password = os.environ.get("GMAIL_APP_PASSWORD", "")
if not gmail_password:
    print("ERROR: GMAIL_APP_PASSWORD not set")
    raise SystemExit(1)

msg = MIMEMultipart("mixed")
msg["Subject"] = f"【模擬操盤報表】{today} 台股收盤"
msg["From"]    = gmail_user
msg["To"]      = gmail_user
msg.attach(MIMEText(report, "plain", "utf-8"))

if HAS_EXCEL:
    xlsx_bytes = make_excel()
    part = MIMEBase("application",
                    "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part.set_payload(xlsx_bytes)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", "attachment",
                    filename=f"台股模擬報表_{today}.xlsx")
    msg.attach(part)
    print(f"Excel 附件已產生（{len(xlsx_bytes):,} bytes）")

with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
    server.login(gmail_user, gmail_password)
    server.send_message(msg)
    print("報表已寄出！")
