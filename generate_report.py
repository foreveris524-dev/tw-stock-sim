"""
台股模擬操盤 — Excel 報表產生器
讀取 portfolio.json / futures_portfolio.json，輸出格式化 .xlsx
"""
import json, os, sys
from datetime import datetime
from urllib.request import urlopen, Request

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("請先安裝 openpyxl：pip install openpyxl")

# ── 色票 ────────────────────────────────────────────────────────────────────
C_DARK_BG   = "1E2329"   # 深色背景
C_HEADER_BG = "2B3139"   # 表頭深灰
C_BLUE      = "3B82F6"   # 強調藍
C_GREEN     = "10B981"   # 正損益
C_RED       = "EF4444"   # 負損益
C_YELLOW    = "F59E0B"   # 警示
C_LIGHT_ROW = "252D36"   # 斑馬淺
C_TEXT      = "E5E7EB"   # 主文字
C_MUTED     = "9CA3AF"   # 次要文字
C_WHITE     = "FFFFFF"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=C_TEXT, size=10):
    return Font(bold=bold, color=color, name="Calibri", size=size)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border():
    s = Side(border_style="thin", color="3A4351")
    return Border(left=s, right=s, top=s, bottom=s)

# ── 資料載入 ─────────────────────────────────────────────────────────────────
def load_json(path):
    if os.path.exists(path):
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    # fallback: GitHub
    TOKEN = os.environ.get("GITHUB_TOKEN", "")
    REPO  = "foreveris524-dev/tw-stock-sim"
    fname = os.path.basename(path)
    import base64
    req = Request(
        f"https://api.github.com/repos/{REPO}/contents/{fname}",
        headers={"Authorization": f"Bearer {TOKEN}", "User-Agent": "ReportGen/1.0"}
    )
    with urlopen(req, timeout=15) as r:
        d = json.loads(r.read())
    return json.loads(base64.b64decode(d["content"]).decode())

BASE = os.path.dirname(__file__)
p  = load_json(os.path.join(BASE, "portfolio.json"))
fp = load_json(os.path.join(BASE, "futures_portfolio.json"))

TODAY = datetime.now().strftime("%Y-%m-%d")
REPORT_DATE = p.get("last_updated", TODAY)[:10]

# ── 衍生數值 ─────────────────────────────────────────────────────────────────
cash         = p.get("cash", 0)
total_value  = p.get("total_value", 0)
holdings     = p.get("holdings", [])
stock_log    = p.get("trade_log", [])

avail_margin = fp.get("available_margin", 0)
used_margin  = fp.get("used_margin", 0)
init_margin  = fp.get("initial_margin", 0)
cum_pnl      = fp.get("cumulative_pnl", 0)
position     = fp.get("position")
fut_log      = fp.get("trade_log", [])

net_assets   = total_value + avail_margin + cum_pnl
initial_cap  = p.get("initial_capital", 1000000) + init_margin
pnl_total    = net_assets - initial_cap
pnl_pct      = pnl_total / initial_cap * 100 if initial_cap else 0

margin_rate  = used_margin / init_margin * 100 if init_margin else 0

# ── Workbook 初始化 ──────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)   # 移除預設空白 sheet

# ════════════════════════════════════════════════════════════════════════════
# Sheet 1：帳戶總覽
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("帳戶總覽")
ws1.sheet_view.showGridLines = False
ws1.sheet_properties.tabColor = C_BLUE

# 設定背景
ws1.sheet_format.defaultRowHeight = 22

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

set_col_widths(ws1, [3, 22, 18, 18, 10])

# 標題列
ws1.merge_cells("B1:E1")
c = ws1["B1"]
c.value = f"📊  台股模擬操盤  —  帳戶總覽  ({REPORT_DATE})"
c.font = Font(bold=True, color=C_WHITE, name="Calibri", size=14)
c.fill = fill(C_DARK_BG)
c.alignment = center()
ws1.row_dimensions[1].height = 36

# 空行
for col in "BCDE":
    ws1[f"{col}2"].fill = fill(C_DARK_BG)

def kv_row(ws, row, label, value, val_color=C_TEXT, note=""):
    ws.row_dimensions[row].height = 24
    for col in "BCDE":
        ws[f"{col}{row}"].fill = fill(C_HEADER_BG if row % 2 == 0 else C_LIGHT_ROW)
    lc = ws[f"B{row}"]
    lc.value = label
    lc.font = font(color=C_MUTED)
    lc.alignment = left()
    lc.border = thin_border()

    vc = ws[f"C{row}"]
    vc.value = value
    vc.font = font(bold=True, color=val_color, size=11)
    vc.alignment = left()
    vc.border = thin_border()

    if note:
        nc = ws[f"D{row}"]
        nc.value = note
        nc.font = font(color=C_MUTED, size=9)
        nc.alignment = left()
        nc.border = thin_border()
    ws[f"E{row}"].fill = fill(C_HEADER_BG if row % 2 == 0 else C_LIGHT_ROW)
    ws[f"E{row}"].border = thin_border()

row = 3
ws1.merge_cells(f"B{row}:E{row}")
sec = ws1[f"B{row}"]
sec.value = "▌ 現貨帳戶"
sec.font = Font(bold=True, color=C_BLUE, name="Calibri", size=11)
sec.fill = fill(C_DARK_BG)
sec.alignment = left()
ws1.row_dimensions[row].height = 28

kv_row(ws1, 4,  "可用現金",     f"${cash:,.0f}")
kv_row(ws1, 5,  "持股市值",     f"${total_value - cash:,.0f}")
kv_row(ws1, 6,  "帳戶總值",     f"${total_value:,.0f}", C_BLUE)
kv_row(ws1, 7,  "持股檔數",     f"{len(holdings)} 檔")

row = 9
ws1.merge_cells(f"B{row}:E{row}")
sec = ws1[f"B{row}"]
sec.value = "▌ 期貨帳戶"
sec.font = Font(bold=True, color=C_YELLOW, name="Calibri", size=11)
sec.fill = fill(C_DARK_BG)
sec.alignment = left()
ws1.row_dimensions[row].height = 28

pos_str = "無持倉"
if position:
    d = position.get("direction","")
    c_ = position.get("contracts",0)
    ep = position.get("entry_price",0)
    pos_str = f"{'空' if d=='SHORT' else '多'} {c_}口 @ {ep:,}"

kv_row(ws1, 10, "目前持倉",      pos_str)
kv_row(ws1, 11, "可用保證金",    f"${avail_margin:,.0f}")
kv_row(ws1, 12, "使用中保證金",  f"${used_margin:,.0f}")
kv_row(ws1, 13, "保證金使用率",  f"{margin_rate:.1f}%",
       C_RED if margin_rate > 70 else C_GREEN)
kv_row(ws1, 14, "累計損益",      f"${cum_pnl:+,.0f}",
       C_GREEN if cum_pnl >= 0 else C_RED)

row = 16
ws1.merge_cells(f"B{row}:E{row}")
sec = ws1[f"B{row}"]
sec.value = "▌ 整體資產"
sec.font = Font(bold=True, color=C_GREEN, name="Calibri", size=11)
sec.fill = fill(C_DARK_BG)
sec.alignment = left()
ws1.row_dimensions[row].height = 28

pnl_color = C_GREEN if pnl_total >= 0 else C_RED
kv_row(ws1, 17, "整體總資產",    f"${net_assets:,.0f}",  C_WHITE)
kv_row(ws1, 18, "初始資本",      f"${initial_cap:,.0f}")
kv_row(ws1, 19, "整體損益",      f"${pnl_total:+,.0f}",  pnl_color)
kv_row(ws1, 20, "損益率",        f"{pnl_pct:+.2f}%",     pnl_color)

# 全背景底色
for r in range(1, 25):
    for col in "ABCDEF":
        cell = ws1[f"{col}{r}"]
        if cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF", ""):
            cell.fill = fill(C_DARK_BG)

# ════════════════════════════════════════════════════════════════════════════
# Sheet 2：現貨操作紀錄
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("現貨操作紀錄")
ws2.sheet_view.showGridLines = False
ws2.sheet_properties.tabColor = C_GREEN
ws2.sheet_format.defaultRowHeight = 20

STOCK_HEADERS = ["日期", "時間", "動作", "代號", "價格", "股數", "金額", "說明"]
COL_WIDTHS_S  = [3, 12, 8, 14, 8, 10, 10, 12, 50]
set_col_widths(ws2, COL_WIDTHS_S)

# 標題
ws2.merge_cells("B1:I1")
t = ws2["B1"]
t.value = "現貨操作紀錄"
t.font = Font(bold=True, color=C_WHITE, name="Calibri", size=13)
t.fill = fill(C_DARK_BG)
t.alignment = center()
ws2.row_dimensions[1].height = 32

# 表頭
for ci, h in enumerate(STOCK_HEADERS, 2):
    cell = ws2.cell(row=2, column=ci)
    cell.value = h
    cell.font = font(bold=True, color=C_WHITE)
    cell.fill = fill(C_BLUE)
    cell.alignment = center()
    cell.border = thin_border()
ws2.row_dimensions[2].height = 24

# 動作色碼
ACTION_COLOR = {
    "BUY":          C_GREEN,
    "SELL":         C_RED,
    "ADJUST_LIMIT": C_YELLOW,
    "PLAN_CANCEL":  C_RED,
    "PLAN_BUY":     "6EE7B7",
    "PLAN_SKIP":    C_MUTED,
    "OBSERVE":      C_MUTED,
    "RESET":        C_MUTED,
}

for ri, entry in enumerate(stock_log, 3):
    bg = C_LIGHT_ROW if ri % 2 == 0 else C_HEADER_BG
    action = entry.get("action", "")
    row_data = [
        entry.get("date", ""),
        entry.get("time", ""),
        action,
        entry.get("code", ""),
        entry.get("price", "") or entry.get("new_limit", ""),
        entry.get("shares", "") or "",
        entry.get("amount", "") or "",
        entry.get("reason", ""),
    ]
    for ci, val in enumerate(row_data, 2):
        cell = ws2.cell(row=ri, column=ci)
        cell.value = val if val != 0 or ci in (5,6,7) else ""
        cell.fill = fill(bg)
        cell.border = thin_border()
        acol = ACTION_COLOR.get(action, C_TEXT)
        if ci == 4:  # 動作欄
            cell.font = font(bold=True, color=acol)
            cell.alignment = center()
        elif ci in (5, 6, 7):
            cell.font = font(color=C_TEXT)
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif ci == 9:
            cell.font = font(color=C_MUTED, size=9)
            cell.alignment = left()
        else:
            cell.font = font(color=C_TEXT)
            cell.alignment = center()
    ws2.row_dimensions[ri].height = 20

# 全背景
for r in range(1, len(stock_log) + 5):
    cell = ws2[f"A{r}"]
    cell.fill = fill(C_DARK_BG)
    cell = ws2[f"J{r}"]
    cell.fill = fill(C_DARK_BG)

# ════════════════════════════════════════════════════════════════════════════
# Sheet 3：期貨操作紀錄
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("期貨操作紀錄")
ws3.sheet_view.showGridLines = False
ws3.sheet_properties.tabColor = C_YELLOW
ws3.sheet_format.defaultRowHeight = 20

FUT_HEADERS = ["日期", "時間", "動作", "方向", "口數", "進場點", "現價", "損益(元)", "說明"]
COL_WIDTHS_F = [3, 12, 8, 14, 8, 8, 10, 10, 12, 50]
set_col_widths(ws3, COL_WIDTHS_F)

ws3.merge_cells("B1:J1")
t = ws3["B1"]
t.value = "期貨操作紀錄（小台指 MTX）"
t.font = Font(bold=True, color=C_WHITE, name="Calibri", size=13)
t.fill = fill(C_DARK_BG)
t.alignment = center()
ws3.row_dimensions[1].height = 32

for ci, h in enumerate(FUT_HEADERS, 2):
    cell = ws3.cell(row=2, column=ci)
    cell.value = h
    cell.font = font(bold=True, color=C_WHITE)
    cell.fill = fill(C_YELLOW)
    cell.alignment = center()
    cell.border = thin_border()
ws3.row_dimensions[2].height = 24

FUT_ACTION_COLOR = {
    "SHORT":        C_RED,
    "LONG":         C_GREEN,
    "CLOSE":        C_BLUE,
    "PLAN_SHORT":   "FCA5A5",
    "PLAN_CANCEL":  C_RED,
    "OBSERVE":      C_MUTED,
    "RESET":        C_MUTED,
}

for ri, entry in enumerate(fut_log, 3):
    bg = C_LIGHT_ROW if ri % 2 == 0 else C_HEADER_BG
    action    = entry.get("action", "")
    direction = entry.get("direction", "")
    pnl       = entry.get("pnl", 0)
    row_data  = [
        entry.get("date", ""),
        entry.get("time", ""),
        action,
        direction or "—",
        entry.get("contracts", 0) or "—",
        entry.get("entry_price", 0) or entry.get("current_price", "—"),
        entry.get("current_price", ""),
        pnl if pnl else "",
        entry.get("reason", ""),
    ]
    for ci, val in enumerate(row_data, 2):
        cell = ws3.cell(row=ri, column=ci)
        cell.value = val
        cell.fill = fill(bg)
        cell.border = thin_border()
        if ci == 4:  # 動作
            acol = FUT_ACTION_COLOR.get(action, C_TEXT)
            cell.font = font(bold=True, color=acol)
            cell.alignment = center()
        elif ci == 5:  # 方向
            dcol = C_RED if direction == "SHORT" else (C_GREEN if direction == "LONG" else C_MUTED)
            cell.font = font(bold=True, color=dcol)
            cell.alignment = center()
        elif ci == 9:  # 損益
            pcol = C_GREEN if (pnl or 0) > 0 else (C_RED if (pnl or 0) < 0 else C_MUTED)
            cell.font = font(bold=True, color=pcol)
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif ci == 10:
            cell.font = font(color=C_MUTED, size=9)
            cell.alignment = left()
        else:
            cell.font = font(color=C_TEXT)
            cell.alignment = center()
    ws3.row_dimensions[ri].height = 20

for r in range(1, len(fut_log) + 5):
    ws3[f"A{r}"].fill = fill(C_DARK_BG)
    ws3[f"K{r}"].fill = fill(C_DARK_BG)

# ── 輸出 ─────────────────────────────────────────────────────────────────────
OUT = os.path.join(os.path.expanduser("~/Desktop"), f"台股模擬報表_{REPORT_DATE}.xlsx")
wb.save(OUT)
print(f"✅ 報表已儲存：{OUT}")
